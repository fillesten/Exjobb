import os.path
import re
from openpyxl import *
from os import path
from openpyxl.utils import column_index_from_string

class ExcelManipulator:
    """
    :param filename: Excel file to open
    """
    def __init__(self, filename):
        if path.exists(filename):
            self.result = os.path.splitext(filename)[0] + " Finished.xlsx"
            self.wb = load_workbook(filename)
        else:
            print("filename not found")
        self.ws = self.wb.active
        self.rowList = self.rowCellData(self.ws)
        self.inputCellData = []
        self.columnParam = None
        self.artikel_beskrivning_info = self.saveColumnToArray(self.ws, 2)

        tradeMarkFile = "ExcelFiles/VarumärkenWL.xlsx"
        if path.exists(tradeMarkFile):
            wlwb = load_workbook(tradeMarkFile)
            wlws = wlwb.active
            self.wlTradeMark = self.saveColumnToArray(wlws, 2)

        self.artikel_nummer_col_index = None
        self.artikel_beskrivning_col_index = None
        self.typbeteckning_col_index = None
        self.kompletterande_info_col_index = None
        self.trademark_col_index = None
        self.write_indices = []

    """ saves all cells Row wise"""
    def rowCellData(self, ws):
        cells = []
        for row in ws.iter_rows():
            rowData = []
            for cell in row:
                if cell.value is not None:
                    cellValue = cell.value
                else:
                    cellValue = ''
                rowData.append(cellValue)
            cells.append(rowData)
        return cells

    """The working function that calls all other functions"""
    def workFunction(self, columnParams=None):
        # converts the given columns no matter the input, for example it takes 1,2 as input as well as 'A','B'
        self.columnWork(columnParams)

        # these all write and manipulate data
        self.writeArtikelnummer()
        self.writeArtikelbeskrivning()
        self.writeTypbeteckning()
        self.writeKompletterandeInfo()
        self.writeTradeMark()
        self.removeDuplicates()

    """save a column from a specific worksheet into an array"""
    def saveColumnToArray(self, ws, col):
        values = []
        for row in ws.iter_rows(min_row=1, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    values.append(cell.value)
        return values

    """A function to set a cells value, it Appends and does NOT replace"""
    def setCellValue(self, row_idx, col_idx, value):
        cell = self.ws.cell(row=row_idx, column=col_idx)
        current_value = cell.value or ''
        new_value = f"{current_value.strip()} {value}" if current_value else str(value)
        cell.value = new_value

    """A function to get a cells value"""
    def getCellValue(self, row_idx, col_idx):
        return self.ws.cell(row=row_idx, column=col_idx).value

    "creates the output columns and also determines the columns that will be iterated through"
    def columnWork(self, columnParam):
        """
           :param columnParam: a list of columns the user wants to iterate through
               can either be numerical or letters, but must be given as a list.
        """
        if columnParam is None: # then iterate over all columns
            columnParam = list(range(1, len(self.rowList[0]) + 1))
        else:
            convertedSearchColumns = []
            for col in columnParam:
                if isinstance(col, str):
                    convertedSearchColumns.append(column_index_from_string(col))
                elif isinstance(col, int):
                    convertedSearchColumns.append(col)
            columnParam = convertedSearchColumns

        self.columnParam = columnParam
        new_columns = ['Artikelnummer', 'Artikelbeskrivning', 'Typbeteckning', 'Kompletterande info', 'Varumärke']
        # + 1 means 1 more empty column
        for idx, header in enumerate(new_columns, len(self.rowList[0]) + 2):
            self.ws.cell(row=1, column=idx).value = header
        self.identifyColindex()
        self.inputRowData()

    """defines the index of the newly created columns and creates an iterable list of the new indices"""
    def identifyColindex(self):
        for idx, cell in enumerate(self.ws[1]):
            if cell.value == "Artikelnummer":
                self.artikel_nummer_col_index = idx + 1
                self.write_indices.append(self.artikel_nummer_col_index)
            if cell.value == "Artikelbeskrivning":
                self.artikel_beskrivning_col_index = idx + 1
                self.write_indices.append(self.artikel_beskrivning_col_index)
            if cell.value == "Typbeteckning":
                self.typbeteckning_col_index = idx + 1
                self.write_indices.append(self.typbeteckning_col_index)
            if cell.value == "Kompletterande info":
                self.kompletterande_info_col_index = idx + 1
                self.write_indices.append(self.kompletterande_info_col_index)
            if cell.value == "Varumärke":
                self.trademark_col_index = idx + 1
                self.write_indices.append(self.trademark_col_index)

    """gives a list of all active cells/columns or the indices of the input columns """
    def inputRowData(self):
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            for col_idx, cell_value in enumerate(cell_row, 1):
                if cell_value is not None and col_idx in self.columnParam:
                    self.inputCellData.append(cell_value)

    """ Change The regular expression variable Artikelnummer to change the desired outcome... """
    def writeArtikelnummer(self):
        # Artikelnummer = r'^\d+$' # accepts only clean digits and in a lone column
        # Artikelnummer = r'^[\d\\/-]+$' #only accepts if the digit is in a lone column
        Artikelnummer = r'\b\d+(?:[-/]\d+)?\b' # accepts digits like 25/3 and 32-53 anywhere
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            for col_idx, cell_value in enumerate(cell_row, 1):
                if col_idx in self.columnParam:  # ensures it only reads from the given columns
                    matches = re.findall(Artikelnummer, str(cell_value))
                    if matches:
                        current_cell_value = self.getCellValue(row_idx, self.artikel_nummer_col_index)
                        longest_match = max(matches, key=lambda x: len(x))
                        if len(longest_match) >= len(str(current_cell_value)):
                            self.ws.cell(row_idx, self.artikel_nummer_col_index).value = ''
                            self.setCellValue(row_idx, self.artikel_nummer_col_index, longest_match)

    """ Can change a variable inside to further clean the Artikelbeskrivning column """
    def writeArtikelbeskrivning(self):
        Artikelbeskrivning = r'\b(?:[A-Za-zåäöÅÄÖ]+(?:[-/][A-Za-zåäöÅÄÖ]+)?)\b'
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            for col_idx, cell_value in enumerate(cell_row, 1):
                if col_idx in self.columnParam:  # ensures it only reads from the given columns
                    matches = re.findall(Artikelbeskrivning, str(cell_value))
                    matches = [match for match in matches if match not in self.wlTradeMark]
                    if matches:
                        self.setCellValue(row_idx=row_idx, col_idx=self.artikel_beskrivning_col_index, value=' '.join(matches))

    """Writes to the Typbeteckning column"""
    def writeTypbeteckning(self):
        # Regular expression to match the pattern
        typbeteckning_pattern = r'\b[A-Za-z0-9\-]*[0-9][A-Za-z0-9\-]*\b'
        # Loop through each row except the header
        for row_number, row in enumerate(self.rowList[1:], 2):
            # Create an iterator for cell values in relevant columns
            cell_values_iterator = ((index + 1, cell) for index, cell in enumerate(row)
                                    if (index + 1) in self.columnParam)
            # Loop through each cell value from the iterator
            for column_number, cell_value in cell_values_iterator:
                # Loop through each word in the cell value
                for word in str(cell_value).split():
                    # Get the current cell value
                    current_cell_value = self.getCellValue(row_number, self.typbeteckning_col_index)
                    # If the word matches the pattern
                    if re.search(typbeteckning_pattern, word):
                        # If the cell is empty, set its value to the word
                        if not current_cell_value:
                            self.setCellValue(row_number, self.typbeteckning_col_index, word)
                        # If the word is longer or equal to the current cell value, replace it
                        elif len(word) >= len(current_cell_value):
                            self.ws.cell(row=row_number, column=self.typbeteckning_col_index).value = ''
                            self.setCellValue(row_number, self.typbeteckning_col_index, word)

    """Writes to the Kompletterande_info column"""
    def writeKompletterandeInfo(self):
        kompletterande_info = r'(\b\w*[^a-zA-ZåäöÅÄÖ\s]\w*\b|\|[^|]+\||<[^>]+>)'
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            for col_idx, cell_value in enumerate(cell_row, 1):
                if col_idx in self.columnParam:  # ensures it only reads from the given columns
                    matches = re.findall(kompletterande_info, str(cell_value))
                    if matches:
                        self.setCellValue(row_idx, self.kompletterande_info_col_index, ' '.join(matches))

    "denna slår emot whitelist, tar första varumärke"
    def writeTradeMark(self):
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            for col_idx, cell_value in enumerate(cell_row, 1):
                if cell_value is not None and col_idx in self.columnParam:
                    for word in str(cell_value).split():
                        if word in self.wlTradeMark:
                            if self.getCellValue(row_idx, self.trademark_col_index) is None:
                                self.setCellValue(row_idx, self.trademark_col_index, word)

    """Checks and removes duplicates in the Kompletterande_info column"""
    def removeDuplicates(self):
        for row_idx, cell_row in enumerate(self.rowList[1:], 2):
            typbeteckning = str(self.getCellValue(row_idx, self.typbeteckning_col_index))
            kompletterande_info = str(self.getCellValue(row_idx, self.kompletterande_info_col_index))
            if typbeteckning in kompletterande_info:
                kompletterande_info = kompletterande_info.replace(typbeteckning, '')
                self.ws.cell(row_idx, self.kompletterande_info_col_index).value = ''
                self.setCellValue(row_idx, self.kompletterande_info_col_index, kompletterande_info)

    """Saves the worksheet and creates the new file"""
    def saveWB(self):
        self.wb.save(self.result)
        self.wb.close()
